import openpyxl
import requests
from typing import List
from pydantic import BaseModel
from langchain.agents import create_agent
from langchain_openai import ChatOpenAI
from langchain_groq import ChatGroq
from django.conf import settings
from dotenv import load_dotenv
import os
load_dotenv()

class PersonalReflection(BaseModel):
    what_i_am_most_proud_of: List[str]
    areas_i_am_focused_on_for_growth: List[str]

class WorkAccomplishments(BaseModel):
    goals_of_this_quarter: List[str]
    goals_of_this_month: List[str]
    official_project_accomplishments: List[str]
    personal_project_accomplishments: List[str]
    personal_reflection: PersonalReflection

class EmployeeInfo(BaseModel):
    work_accomplishments: WorkAccomplishments
    learning: List[str]
    utilized_skills: List[str]

def extract_employee_data(file_path: str, employee_name: str, month: str, column_name: str = 'What I did  yesterday?') -> List[str]:
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        found_values = []

        for sheet_name in workbook.sheetnames:
            if month.lower() not in sheet_name.lower():
                continue

            sheet = workbook[sheet_name]
            header = [cell.value for cell in sheet[1]]
            
            try:
                col_index = header.index(column_name)
            except ValueError:
                continue

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if employee_name in row:
                    value = row[col_index]
                    if value:
                        found_values.append(value)

        return found_values
    except Exception as e:
        raise Exception(f"Error extracting data: {str(e)}")

def get_available_models(provider: str, api_key: str) -> List[dict]:
    """Get available models for the specified provider"""
    try:
        if provider == 'gemini':
            # Gemini API models endpoint
            url = "https://generativelanguage.googleapis.com/v1beta/models"
            headers = {"x-goog-api-key": api_key}
            
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            
            data = response.json()
            models = []
            
            for model in data.get('models', []):
                if 'generateContent' in model.get('supportedGenerationMethods', []):
                    model_id = model['name'].split('/')[-1]
                    models.append({
                        'id': model_id,
                        'name': model.get('displayName', model_id)
                    })
            
            return models
            
        elif provider == 'groq':
            # Groq API models endpoint
            url = "https://api.groq.com/openai/v1/models"
            headers = {"Authorization": f"Bearer {api_key}"}
            
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            
            data = response.json()
            models = []
            print("groq response models data: \n", data)
            for model in data.get('data', []):
                models.append({
                    'id': model['id'],
                    'name': model['id']
                })
            
            return models
        
        elif provider == 'ollama':
            url = "http://localhost:11434/api/tags"
            headers = {"Authorization": f"Bearer {api_key}"}
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            data = response.json()
            print("ollama response models data: \n", data)
            models = []
            
            for model in data.get('models', []):
                models.append({
                    'id': model['model'],
                    'name': model['name']
                })
            
            return models

            
        else:
            raise ValueError(f"Unsupported provider: {provider}")
            
    except Exception as e:
        raise Exception(f"Error fetching models for {provider}: {str(e)}")

def generate_brag_document(employee_data: List[str], provider: str, api_key: str, model: str) -> EmployeeInfo:
    # Initialize the appropriate LLM based on provider
    if provider == 'gemini':
        llm = ChatOpenAI(
            model=model,
            base_url=os.getenv("GEMINI_API_BASE_URL"),
            api_key=api_key,
        )
    elif provider == 'groq':
        llm = ChatOpenAI(
            model=model,
            api_key=api_key,
            base_url = os.getenv("GROQ_API_BASE_URL")
            
        )
    elif provider == 'ollama':
        llm = ChatOpenAI(
            model=model,
            api_key=api_key,
            base_url = os.getenv("OLLAMA_API_BASE_URL")
        )
    else:
        raise ValueError(f"Unsupported provider: {provider}")
    
    agent = create_agent(
        model=llm,
        tools=[],
        response_format=EmployeeInfo,
        system_prompt="""You are an AI assistant that creates professional brag documents from raw input notes, emphasizing achievements, learning, and utilized skills. Follow these instructions:

**Instructions:**

1. **Input:** Raw notes about accomplishments, learning, and skills.
2. **Output:** A brag document in three sections:
3. **Caution:** Never ever add any irrelavant information which are not in the working history. Genearate only based on the provided employee information.

**Output Structure:**

* **WORK ACCOMPLISHMENTS**
  * Focus on specific results, improvements, or contributions.
  * Use past tense and action-oriented language.
  * Do not add ticket number or Task number (e.g. KR-619 etc.).
  There are five subcategories such as 'Goals of this Quarter', 'Goals of this Month', 'Official Project Accomplishments', 'Personal Project Accomplishments' and 'Personal Reflection'. Inside 'Personal Reflection', there are two subcategories such as 'What I'm Most Proud Of' and 'Areas I'm Focused On For Growth'.



* **LEARNING**

  * Convert raw notes about courses, books, or self-study into bullet points.
  * Highlight practical knowledge, skills gained, or tools explored.

* **UTILIZED SKILLS**

  * Extract technologies, frameworks, or tools used.
  * Present as a comma-separated list or bullets.

**Formatting Rules:**

* Do **not** include a title/header; start with the first section.
* Every line must be a single bullet point.
* Keep language professional, precise, and specific.
* Avoid vague statements; emphasize measurable or observable impact.
* Do not add more than 10 bullet points for any section. Do not add more than 7 skills.
* Use section headers exactly as: `WORK ACCOMPLISHMENTS`, `LEARNING`, `UTILIZED SKILLS`.

    """
    )
    
    response = agent.invoke(
        {"messages": [{"role": "user", "content": f"Here is the employee information: {' '.join(employee_data)}"}]}
    )
    
    return response["structured_response"]
